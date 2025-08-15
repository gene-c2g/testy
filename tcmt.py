##################################################
## tcmt.py
##################################################
## Author: gene@cloud2gnd.com
## Copyright: Copyright 2025
## Version: 1.0
##################################################

import docx.table
import docx.text
import docx.text.paragraph
from bcolors import bcolors
from wordhelper import getAcceptedText
from excelhelper import setHeaderCell

from docx import Document
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
from docx.enum.section import WD_ORIENT,WD_SECTION
from docx.enum.dml import MSO_COLOR_TYPE

import openpyxl
from openpyxl.cell.text import InlineFont
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.cell.rich_text import TextBlock, CellRichText
import docx
import os
import re
import nltk


################################################################################################################################################

def extractTCMT(filename):
    print("extractTS:", filename)
    if  not os.path.exists(filename):
        print(bcolors.FAIL + "File '%s' does not exist" % filename + bcolors.ENDC)
        return
    tcmtEntries = []

    print("Extracing TCMT from %s" % filename)
    parseTestDocument(filename, lambda x, y: tcmtEntries.append((x, y)) )

    outputXLS(filename, tcmtEntries)
    #if filename.find(".tex") != -1:
    #    outputFile = filename.le
   # outputTeX(filename, testText)

def removeBrace(line):
    return line.replace("}", "")
        
def outputXLS(filename, tcmtEntries):
    (root,ext) = os.path.splitext(filename)
    newfile = root+"_TCA.xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.sheet_view.zoomScale = 140
    setHeaderCell(sheet, 1, "ICS", 20)
    setHeaderCell(sheet, 2, "TCIDs", 50)
    nRow = 2
    for tcmtEntry in tcmtEntries:
        sheet.cell(row = nRow, column = 1).value = removeBrace(tcmtEntry[0])
        sheet.cell(row = nRow, column = 2).value = removeBrace(tcmtEntry[1])
        nRow += 1
    wb.save(newfile)
    print("Created %s" % newfile)

def hasMatchingBraces(text):
    stack = []
    matches = []

    for i, char in enumerate(text):
        if char == '{':
            stack.append(i)
        elif char == '}':
            if stack:
                start = stack.pop()
                matches.append(text[start:i+1])
            else:
                # Unmatched closing brace
                print(f"Unmatched closing brace at position {i}")

    if stack:
        for unmatched in stack:
            print(f"Unmatched opening brace at position {unmatched}")
            return False

    return True


def parseTestDocument(filename, adddelegate):
    tcmtEntry = [None] * 2
    completeLine = ""
    lineCompleted = False
    inTableCell = False
    with open(filename) as file:
        for line in file:
            #print (line)
            if inTableCell:
                completeLine += " " + (line)
            else:
                match = re.search(r"TableCell(.*)", line)
                if(match):
                    inTableCell = True
                    substr = match.group(1).strip()
                    completeLine += substr

            if len(completeLine)> 0 and hasMatchingBraces(completeLine):
                lineCompleted = True
            if(lineCompleted):
                inTableCell = False
                lineCompleted = False
                match = re.search (r"\\TCMTICS{(.*)}", completeLine)
                if match:
                    print ("found: ", match.group(1))
                    ics = match.group(1)
                    tcmtEntry[0] = match.group(1)
                match = re.search (r"\\TCMTTCID{(.*)}", completeLine)
                if match:
                    print ("found: ", match.group(1))
                    tcmt = match.group(1)
                    tcmtEntry[1] = match.group(1)
                completeLine = ""
                if type(tcmtEntry[0]) != type(None) and type(tcmtEntry[1]) != type(None) and len(tcmtEntry[0]) and len(tcmtEntry[1]):
                    adddelegate(tcmtEntry[0], tcmtEntry[1])
                    tcmtEntry = [None] * 2
    return

        